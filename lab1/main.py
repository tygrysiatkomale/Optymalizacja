import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import random
import user_func
import algorithms


def make_plot(f, range_start=-100, range_stop=100):
    x = []
    y = []

    for i in range(range_start, range_stop):
        x.append(i)
        y.append(f(i))

    plt.plot(x, y, color='b', label='Wartości')

    plt.title('Wykres funkcji')
    plt.xlabel('Oś X')
    plt.ylabel('Oś Y')
    plt.legend()
    plt.grid(True)
    plt.show()
    return 0


# dodawanie wartosci do arkusza "symulacja"
def add_to_excel_simulation_results(DA, sol, method):
    file_name = "xlsx1.xlsx"
    try:
        workbook = load_workbook(file_name)
    except FileNotFoundError:
        workbook = Workbook()

    if 'Symulacja' in workbook.sheetnames:
        sheet = workbook['Symulacja']
    else:
        sheet = workbook.create_sheet('Symulacja')

    # Ustal kolumny dla odpowiednich danych
    col_map = {
        'Fibonacci': {'VA': 'B', 'VB': 'D', 'TB': 'F'},
        'Lagrange': {'VA': 'C', 'VB': 'E', 'TB': 'G'}
    }

    # Znajdź pierwszy wolny wiersz poniżej nagłówków (pomijamy połączone komórki, zaczynamy od trzeciego wiersza)
    start_row = 3
    row = sheet.max_row + 1 if sheet.max_row >= start_row else start_row

    # Dodajemy dane symulacyjne zgodnie z istniejącymi nagłówkami
    for i, t in enumerate(sol.t):
        current_row = row + i

        # Sprawdzamy, czy wiersz istnieje, a jeśli nie, tworzymy go
        if sheet[f'A{current_row}'].value is None:
            sheet[f'A{current_row}'] = t

        # Wpisujemy dane do odpowiednich kolumn dla VA, VB, TB
        if method in col_map:
            if sheet[f"{col_map[method]['VA']}{current_row}"].value is None:
                sheet[f"{col_map[method]['VA']}{current_row}"] = sol.y[0][i]  # VA
            if sheet[f"{col_map[method]['VB']}{current_row}"].value is None:
                sheet[f"{col_map[method]['VB']}{current_row}"] = sol.y[1][i]  # VB
            if sheet[f"{col_map[method]['TB']}{current_row}"].value is None:
                sheet[f"{col_map[method]['TB']}{current_row}"] = sol.y[2][i]  # TB

    # Zapisz zmiany w pliku
    workbook.save(file_name)
    print(f"Wyniki symulacji zostały zapisane do arkusza 'Symulacja' w pliku {file_name}")

    # Printowanie wyników do debugowania
    print(f"Symulacja dla DA = {DA * 10000:.2f} cm^2 ({method}):")
    for i, t in enumerate(sol.t):
        print(f"t = {t:.2f}s, VA = {sol.y[0][i]:.5f} m^3, VB = {sol.y[1][i]:.5f} m^3, TB = {sol.y[2][i]:.2f} °C")


def add_to_excel(x0, a, b, n, fib_ncalls, fib_min, lagrange_ncalls, lagrange_min, line):
    file_name = "xlsx1.xlsx"
    try:
        # Próba załadowania istniejącego pliku
        workbook = load_workbook(file_name)
    except FileNotFoundError:
        # Jeśli plik nie istnieje, tworzymy nowy
        workbook = Workbook()

    # Wybieramy aktywny arkusz (pierwszy z dostępnych)
    sheet = workbook["Tabela 1"]

    sheet[f'C{line}'] = x0
    sheet[f'D{line}'] = a
    sheet[f'E{line}'] = b
    sheet[f'F{line}'] = n
    sheet[f'G{line}'] = fib_min
    sheet[f'H{line}'] = user_func.ff1T(fib_min)
    sheet[f'I{line}'] = fib_ncalls
    sheet[f'J{line}'] = "lokalne" if user_func.ff1T(fib_min) > -0.1 else "globalne"
    sheet[f'K{line}'] = lagrange_min
    sheet[f'L{line}'] = user_func.ff1T(lagrange_min)
    sheet[f'M{line}'] = lagrange_ncalls
    sheet[f'N{line}'] = "lokalne" if user_func.ff1T(lagrange_min) > -0.1 else "globalne"

    # Zapisujemy zmiany w pliku
    workbook.save(file_name)
    # print(f"Wartości zostały zapisane do pliku {file_name} w linii {line}")


# Parametry i testy
x0 = random.randint(-50, 100)
d = 1
alpha = 1.5
nmax = 10000
epsilon = 0.00001

make_plot(user_func.ff1T)

for i in range(3, 303):
    if 3 <= i <= 102:
        alpha = 1.5
    elif 103 <= i <= 202:
        alpha = 2
    elif 203 <= i <= 302:
        alpha = 2.5
    x0 = random.randint(-50, 100)
    expansion_result = algorithms.expansion_method(user_func.ff1T, x0, d, alpha, nmax)
    a, b, fcalls = expansion_result
    fib_result = algorithms.fibonacci_method(user_func.ff1T, a, b, epsilon)
    # c = (a + b) / 2
    c = fib_result[0]
    lagrange_result = algorithms.lagrange_interpolation(user_func.ff1T, a, b, c, epsilon)
    add_to_excel(x0, a, b, fcalls, fib_result[1], fib_result[0], lagrange_result[1], lagrange_result[0], i)

alpha = 1.5
# Symulacja dla Da = 50 cm^2
DA_test = 0.005   # Da w cm^2
difference, max_TB, sol = user_func.ff2R(DA_test)

# # Wyniki dla DA = 50 cm^2
print(f" Maksymalna temperatura w zbiorniku B dla DA = 50 cm^2: {max_TB:.2f}°C")

x1 = 0.005
d1 = 0.002
a1 = 1/10000
b1 = 100/10000
# Metoda ekspansji - wstępne oszacowanie przedziału poszukiwań
a, b, fcalls_expansion = algorithms.expansion_method(lambda DA: user_func.ff2R(DA)[0], x1, d1, alpha)
print(f"Przedział po ekspansji: a = {a:.5f}, b = {b:.5f}, liczba wywołań funkcji celu: {fcalls_expansion}")

# Metoda Fibonacciego - optymalizacja w przedziale [a1, b1]
DA_fib, fcalls_fib = algorithms.fibonacci_method(lambda DA: user_func.ff2R(DA)[0], a1, b1, epsilon)
print(f"Optymalna wartość DA (metoda Fibonacciego): {DA_fib * 10000:.2f} cm^2, "
      f"liczba wywołań funkcji celu: {fcalls_fib}")

# Metoda interpolacji Lagrange'a - optymalizacja w przedziale [a1, b1]
c = (a + b) / 2  # Punkt wewnętrzny dla metody Lagrange'a
try:
    DA_lagrange, fcalls_lagrange = algorithms.lagrange_interpolation(lambda DA: user_func.ff2R(DA)[0],
                                                                     a1, b1, c, epsilon)
    print(f"Optymalna wartość DA (interpolacja Lagrange'a): {DA_lagrange * 10000:.2f} cm^2, "
          f"liczba wywołań funkcji celu: {fcalls_lagrange}")
except ValueError as e:
    print(f"Wystąpił błąd podczas optymalizacji metodą Lagrange'a: {e}")
    DA_lagrange = None

# Wywołanie funkcji celu dla uzyskanego wyniku Fibonacciego, aby zweryfikować temperaturę
_, max_TB_fib, sol_fib = user_func.ff2R(DA_fib)
print(f"Maksymalna temperatura w zbiorniku B dla DA (Fibonacci) = {DA_fib * 10000:.2f} cm^2: {max_TB_fib:.2f}°C")
add_to_excel_simulation_results(DA_fib, sol_fib, "Fibonacci")

# wywołanie funkcji celu dla uzyskanego wyniku lagrange, aby zweryfikowac temperature
if DA_lagrange is not None:
    _, max_TB_lag, sol_lag = user_func.ff2R(DA_lagrange)
    print(f"Maksymalna temperatura w zbiorniku B dla DA (Lagrange) = {DA_lagrange * 10000:.2f} cm^2: "
          f"{max_TB_lag:.2f}°C")
    add_to_excel_simulation_results(DA_lagrange, sol_lag, "Lagrange")


# Wizualizacja wyników symulacji
plt.plot(sol_fib.t, sol_fib.y[1], label='Objętość w zbiorniku B (Fibonacci)', color='b')
plt.plot(sol_fib.t, sol_fib.y[2], label='Temperatura w zbiorniku B (Fibonacci)', color='r')
plt.xlabel('Czas (s)')
plt.ylabel('Wartość')
plt.title('Zmiany w zbiorniku B w czasie (optymalne DA - Fibonacci)')
plt.legend()
plt.grid(True)
plt.show()

if DA_lagrange is not None:
    plt.plot(sol_lag.t, sol_lag.y[1], label='Objętość w zbiorniku B (Lagrange)', color='b')
    plt.plot(sol_lag.t, sol_lag.y[2], label='Temperatura w zbiorniku B (Lagrange)', color='r')
    plt.xlabel('Czas (s)')
    plt.ylabel('Wartość')
    plt.title('Zmiany w zbiorniku B w czasie (optymalne DA - Lagrange)')
    plt.legend()
    plt.grid(True)
    plt.show()
