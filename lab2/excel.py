from turtledemo.chaos import line

from openpyxl import Workbook, load_workbook


def add_to_excel(start_x1, start_x2, hookey_x1, hookey_x2, hookey_y, hookey_count,
                 rosen_x1, rosen_x2, rosen_y, rosen_count, line):
    file_name = "xlsx2.xlsx"
    try:
        # Próba załadowania istniejącego pliku
        workbook = load_workbook(file_name)
    except FileNotFoundError:
        # Jeśli plik nie istnieje, tworzymy nowy
        workbook = Workbook()

    # Wybieramy aktywny arkusz (pierwszy z dostępnych)
    sheet = workbook["Tabela 1"]

    sheet[f'C{line}'] = start_x1
    sheet[f'D{line}'] = start_x2
    sheet[f'E{line}'] = hookey_x1
    sheet[f'F{line}'] = hookey_x2
    sheet[f'G{line}'] = hookey_y
    sheet[f'H{line}'] = hookey_count
    sheet[f'I{line}'] = "Globalne" if hookey_y < 0.01 else "Lokalne"
    sheet[f'J{line}'] = rosen_x1
    sheet[f'K{line}'] = rosen_x2
    sheet[f'L{line}'] = rosen_y
    sheet[f'M{line}'] = rosen_count
    sheet[f'N{line}'] = "Globalne" if rosen_y < 0.01 else "Lokalne"

    # Zapisujemy zmiany w pliku
    workbook.save(file_name)
    # print(f"Wartości zostały zapisane do pliku {file_name} w linii {line}")


def add_to_chart_excel_hooke(x1, x2, line):
    file_name = "xlsx2.xlsx"
    workbook = load_workbook(file_name)
    sheet = workbook["Wykres"]
    sheet[f'B{line}'] = x1
    sheet[f'C{line}'] = x2
    workbook.save(file_name)


def add_to_chart_excel_rosen(x1, x2, line):
    file_name = "xlsx2.xlsx"
    workbook = load_workbook(file_name)
    sheet = workbook["Wykres"]
    sheet[f'D{line}'] = x1
    sheet[f'E{line}'] = x2
    workbook.save(file_name)


def add_to_sym(column_a, column_b, alpfa, omega, line):
    file_name = "xlsx2.xlsx"
    workbook = load_workbook(file_name)
    sheet = workbook["Symulacja"]
    sheet[f'{column_a}{line}'] = alpfa
    sheet[f'{column_b}{line}'] = omega
    workbook.save(file_name)
