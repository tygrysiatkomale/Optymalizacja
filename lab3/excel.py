from turtledemo.chaos import line

from openpyxl import Workbook, load_workbook


def add_to_excel(x1, x2,
                 x1_zew, x2_zew, r_zew, y_zew, n_zew,
                 x1_wew, x2_wew, r_wew, y_wew, n_wew,
                 line):
    file_name = "xlsx2.xlsx"
    try:
        # Próba załadowania istniejącego pliku
        workbook = load_workbook(file_name)
    except FileNotFoundError:
        # Jeśli plik nie istnieje, tworzymy nowy
        workbook = Workbook()

    # Wybieramy aktywny arkusz (pierwszy z dostępnych)
    sheet = workbook["Tabela 1"]

    sheet[f'C{line}'] = x1
    sheet[f'D{line}'] = x2
    sheet[f'E{line}'] = x1_zew
    sheet[f'F{line}'] = x2_zew
    sheet[f'G{line}'] = r_zew
    sheet[f'H{line}'] = y_zew
    sheet[f'I{line}'] = n_zew
    sheet[f'J{line}'] = x1_wew
    sheet[f'K{line}'] = x2_wew
    sheet[f'L{line}'] = r_wew
    sheet[f'M{line}'] = y_wew
    sheet[f'N{line}'] = n_wew

    # Zapisujemy zmiany w pliku
    workbook.save(file_name)
    # print(f"Wartości zostały zapisane do pliku {file_name} w linii {line}")